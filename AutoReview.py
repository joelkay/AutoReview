import openpyxl, pprint, warnings, time, datetime
import FileDialog#for pyinstaller (also make sure setup tools = 19.2)
import dateutil.relativedelta
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from Tkinter import Tk
from tkFileDialog import askopenfilename
warnings.simplefilter("ignore")

######DEBUG#############
debug = 0	
########################
start_time = time.time()
#####################################LOAD A PURCHASE SHEET FROM FOLDER#######################################################

##Load a purchase sheet##
Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

print('Loading'+' '+filename)

wb = openpyxl.load_workbook(filename, data_only = True)#only want values not formulas
wb2 = openpyxl.load_workbook(filename)#open the wb we will save(containing formulas and vba)

sheet = wb.get_sheet_by_name('Main')
sheet2 = wb2.get_sheet_by_name('Main')

print "Workbook contains {rows} number of rows".format(rows=sheet.max_row)

##########################################COLUMN INDEX OF STOCK CODES####################################
ABC=5
NetStock=68 #(col.67) 
OpenDemand=67#(col.66) 
DaysCovered=57#(col,56)
OrderQty=58
Action=59
Comments=60
StockCode=1
Supplier=8
SafetyStock=8
EBQ=11
OnHand=24
P12 = 23#sales value 12 months ago
P11 = 22#sales value 11 months ago
P10 = 21#sales value 10 months ago
P9 = 20#sales value 9 months ago
P8 = 19#sales value 8 months ago
P7 = 18#sales value 7 months ago
P6 = 17#sales value 6 months ago
P5 = 16#sales value 5 months ago
P4 = 15#sales value 4 months ago
P3 = 14#sales value 3 months ago
P2 = 13#sales value 2 months ago
P1 = 12#sales value 1 months ago
P0 = 46#current month WIP
##########################################END OFCOLUMN INDEX OF STOCK CODES####################################
def returnmonth(x):	
	d = datetime.datetime.strptime(time.strftime("%d/%m/%Y"),"%d/%m/%Y")
	d2 = d - dateutil.relativedelta.relativedelta(months=x)
	return d2

#capture last 12 dates as date_time objects
dates = [returnmonth(12),returnmonth(11),returnmonth(10),
		 returnmonth(9),returnmonth(8),returnmonth(7),
		 returnmonth(6),returnmonth(5),returnmonth(4),
		 returnmonth(3),returnmonth(2),returnmonth(1),
		 returnmonth(0),returnmonth(-1),returnmonth(-2),
		 returnmonth(-3)]
#################################################################################################################
######################################## RETURN FORECASTED VALUE#################################################
def returnvalue(rowNum):	
#create an array of dates vs sales       
		df = pd.DataFrame([sheet.cell(row=rowNum, column=P12).value,sheet.cell(row=rowNum, column=P11).value,
						   sheet.cell(row=rowNum, column=P10).value,sheet.cell(row=rowNum, column=P9).value,
						   sheet.cell(row=rowNum, column=P8).value,sheet.cell(row=rowNum, column=P7).value,
						   sheet.cell(row=rowNum, column=P6).value,sheet.cell(row=rowNum, column=P5).value,
						   sheet.cell(row=rowNum, column=P4).value,sheet.cell(row=rowNum, column=P3).value,
						   sheet.cell(row=rowNum, column=P2).value,sheet.cell(row=rowNum, column=P1).value,
						   sheet.cell(row=rowNum, column=P0).value,0,0,0], index = dates, columns=['Actual Sales'] )#import last 12 months sales into an array

		#CURRENT MONTH
		if df['Actual Sales'].iloc[12] < df['Actual Sales'].iloc[1]: df['Actual Sales'].iloc[12] = df['Actual Sales'].iloc[1]
			#if WIP is less than sales last year then use last year's figures instead for our estimation
		#MONTH +1
		df['Actual Sales'].iloc[13] = df['Actual Sales'].iloc[2] #month + 1 = prev value at that point in time
		#MONTH +2
		df['Actual Sales'].iloc[14] = df['Actual Sales'].iloc[3] #month + 2 = prev value at that point in time
		#MONTH +3
		df['Actual Sales'].iloc[15] = df['Actual Sales'].iloc[4] #month + 3 = prev value at that point in time

		df['EWMA'] = pd.ewma(df['Actual Sales'], span = 13)#Forecast Values using EWMA(estimated weight moving average)

		##ADJUST VALUES ACORDING TO THE EWMA
		df['Actual Sales'].iloc[12]  = df['EWMA'].iloc[12]

		df['Actual Sales'].iloc[13] = df['EWMA'].iloc[13]  #month + 1 = value estimated by EWMA algorithm at this point in time
		#MONTH +2
		df['Actual Sales'].iloc[14] = df['EWMA'].iloc[14] #month + 2 = value estimated by EWMA algorithm at this point in time
		#MONTH +3
		df['Actual Sales'].iloc[15] = df['EWMA'].iloc[15] #month + 3 = value estimated by EWMA algorithm at this point in time

		predictednet = sheet.cell(row=rowNum, column=NetStock).value - (df['Actual Sales'].iloc[13] + df['Actual Sales'].iloc[14]) #Current Net Stock - Month 1 and Month 2 demand

		if debug: df.plot(title = sheet.cell(row=rowNum, column=StockCode).value)

		#find out averages/forecasts e.t.c
		#work out ammount to order
		#update sheet
		

		return predictednet #how much stock would we have if we sold this and next months stock(  helps us estimate how much we need to replen  )
#################################################################GET AMMOUNT#################################################################################
def getammount(safety,predictednet,rowNum):
		qty = safety-predictednet#bias field to tweak how much true qty we need. 
		ebq = sheet.cell(row=rowNum, column=EBQ).value 
		temp = ebq#temp value to adjust

		if debug: print str(qty)+' '+str(temp)

		while True:
			if qty > temp:
				if temp == 1:
					temp = 2 #if ebq = 1 make temp = 2
				else:
					temp = temp * 2#check if its the next size up(multiples of 2)
			else:
				if debug: print str(temp)
				qty = temp

				###########################
				if qty > safety * 2:
					qty = safety * 2
				#####health check on large ebqs

				return qty#ammount to order tuned to match ebq

#############################################################################################################################################################



for rowNum in range(2,sheet.max_row+1): # skip the first row,include last row


########################################BLANKS/SPECIAL STOCK CODES######################################
# clear items with No demand and No stock from our attention
	if  sheet.cell(row=rowNum, column=StockCode).value == 'Q8-6296' or sheet.cell(row=rowNum, column=StockCode).value == 'Q8-6296-1':#Direct ship orders
		if debug: print "[Phase:1]Item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'Direct Ship' # JK Comment(col.59) 

	#packaging supplies
	if  sheet.cell(row=rowNum, column=Supplier).value == 'UKWER01' or sheet.cell(row=rowNum, column=Supplier).value == 'UKTHU01' or sheet.cell(row=rowNum, column=Supplier).value == 'UKCRU01':#WER Roberts stuff on Kanban
		if debug: print "[Phase:1]Item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'Packaging Supplies - Kanban' # JK Comment(col.59) 


	#GLOBAL CLEAR OUT CODES THAT are not selling but need to get reset to safety stock.
	if sheet.cell(row=rowNum, column=OpenDemand).value == 0 and sheet.cell(row=rowNum, column=OnHand).value == 0 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if sheet.cell(row=rowNum, column=SafetyStock).value == 0 or sheet.cell(row=rowNum, column=SafetyStock).value == 'NULL':
			if debug: print "[Phase:2]Item marked with No Action"
			sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'No Demand and S/S set at 0' # JK Comment(col.59) 
		else:
			if debug: print "[Phase:2]Item marked with Order Action"
			sheet2.cell(row=rowNum, column=Action).value = 'Order' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'Replen S/Stock' # JK Comment(col.59) 
			sheet2.cell(row=rowNum, column=OrderQty).value = sheet.cell(row=rowNum, column=SafetyStock).value # Order S/Stock Value if it is low.

# mark blank codes as Kit(No Action)
	if sheet.cell(row=rowNum, column=ABC).value is None and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    # Net Stock(col.67) and Open Demand(col.66) is 0
		if debug: print "[Phase:1b]BLANK ABC item marked with Kit(No Action)"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'No ABC field' # JK Comment(col.59) 

############################################HARDWARE#################################
# clear hardware/winch codes from our attention
	if sheet.cell(row=rowNum, column=ABC).value == 'HDW' or sheet.cell(row=rowNum, column=ABC).value == 'WIN' and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    # Net Stock(col.67) and Open Demand(col.66) is 0
		if debug: print "[Phase:1c]HDW item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'Hardware' # JK Comment(col.59) 

##############################################NEW ITEMS###############################
# mark new codes for Phil B's attention
	if sheet.cell(row=rowNum, column=ABC).value == 'NEW' and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    # Net Stock(col.67) and Open Demand(col.66) is 0
		if debug: print "[Phase:1d]NEW item marked with PB to Review"
		sheet2.cell(row=rowNum, column=Action).value = 'PB to Review' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'New Item' # JK Comment(col.59) 


#########################################OBSOLETE###################################
# clear obsolete items from our attention
			
#mark obsolete items with >90 days of stock as No Action
	if sheet.cell(row=rowNum, column=ABC).value == 'OBS' and sheet.cell(row=rowNum, column=DaysCovered).value >= 90 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:2b]Obsolete item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = '>90 days of stock available' # JK Comment(col.59) 

#obsolete that need forecasted information for a decision
	if sheet.cell(row=rowNum, column=ABC).value == 'OBS' and sheet2.cell(row=rowNum, column=Action).value is None and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 

		predictednet = returnvalue(rowNum)
		safety = sheet.cell(row=rowNum, column=SafetyStock).value
		bias = 0# weighted decision

		if predictednet > safety:
			if debug: print "[Phase:2c]Obsolete item marked with No Action"
			sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'Enought S/S to Cover demand' # JK Comment(col.59) 
		else:	
			if debug: print "[Phase:2d]Obsolete item marked with Order Action"
			sheet2.cell(row=rowNum, column=Action).value = 'Order' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=OrderQty).value = getammount(safety,predictednet,rowNum) #Replen safety while ordering to meet predicted demand according to EBQ: bias = 0 default
			sheet2.cell(row=rowNum, column=Comments).value = 'S/S:'+str(safety)+' p.net:'+str(round(predictednet))+' (safety-predicted)= '+str(round(safety-predictednet)) # JK Comment(col.59) 



#########################################STRANGERS###################################
# clear strangers from our attention
#mark stranger items with >90 days of stock as No Action
	if sheet.cell(row=rowNum, column=ABC).value == 'STA' and sheet.cell(row=rowNum, column=DaysCovered).value >= 90 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:3b]Stranger item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = '>90 days of stock available' # JK Comment(col.59) 
#mark stranger items with no safety stock and >=0 net stock as No Action
	if sheet.cell(row=rowNum, column=ABC).value == 'STA' and sheet.cell(row=rowNum, column=SafetyStock).value == 0 and sheet.cell(row=rowNum, column=NetStock).value >= 0 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:3c]Stranger item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = 'No Safety Stock' # JK Comment(col.59) 
#mark stranger items with net stock >= net stock as No Action
	if sheet.cell(row=rowNum, column=ABC).value == 'STA' and sheet.cell(row=rowNum, column=NetStock).value >= sheet.cell(row=rowNum, column=SafetyStock).value and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:3d]Stranger item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = '*Have safety stock on Hand' # JK Comment(col.59) 

#strangers that need forecasted information for a decision
	if sheet.cell(row=rowNum, column=ABC).value == 'STA' and sheet2.cell(row=rowNum, column=Action).value is None and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 

		predictednet = returnvalue(rowNum)
		safety = sheet.cell(row=rowNum, column=SafetyStock).value
		bias = 0# weighted decision

		if predictednet > safety:
			if debug: print "[Phase:3e]Stranger item marked with No Action"
			sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'Enought S/S to Cover demand' # JK Comment(col.59) 
		else:	
			if debug: print "[Phase:3f]Stranger item marked with Order Action"
			sheet2.cell(row=rowNum, column=Action).value = 'Order' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=OrderQty).value = getammount(safety,predictednet,rowNum) #Replen safety while ordering to meet predicted demand according to EBQ: bias = 0 default
			sheet2.cell(row=rowNum, column=Comments).value = 'S/S:'+str(safety)+' p.net:'+str(round(predictednet))+' (safety-predicted) = '+str(round(safety-predictednet)) # JK Comment(col.59) 


#########################################REPEATERS###################################
# clear strangers from our attention
	
#mark repeater items with >90 days of stock as No Action ##DO HEALTH CHECK AGAINST FORECASTS##
	if sheet.cell(row=rowNum, column=ABC).value == 'REP' and sheet.cell(row=rowNum, column=DaysCovered).value >= 90 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:4b]Repeater item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = '>90 days of stock available' # JK Comment(col.59) 

#repeaters that need forecasted information for a decision
	if sheet.cell(row=rowNum, column=ABC).value == 'REP' and sheet2.cell(row=rowNum, column=Action).value is None and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 

		predictednet = returnvalue(rowNum)
		safety = sheet.cell(row=rowNum, column=SafetyStock).value
		bias = 0# weighted decision

		if predictednet > safety:
			if debug: print "[Phase:4c]Repeater item marked with No Action"
			sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'Enought S/S to Cover demand' # JK Comment(col.59) 
		else:	
			if debug: print "[Phase:4d]Repeater item marked with Order Action"
			sheet2.cell(row=rowNum, column=Action).value = 'Order' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=OrderQty).value = getammount(safety,predictednet,rowNum) #Replen safety while ordering to meet predicted demand according to EBQ: bias = 0 default
			sheet2.cell(row=rowNum, column=Comments).value = 'S/S:'+str(safety)+' p.net:'+str(round(predictednet))+' (safety-predicted) = '+str(round(safety-predictednet)) # JK Comment(col.59) 

		
#########################################RUNNERS###################################
# clear runners from our attention
	
#mark runner items with >180 days of stock as No Action##
	if sheet.cell(row=rowNum, column=ABC).value == 'RUN' and sheet.cell(row=rowNum, column=DaysCovered).value >= 180 and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 
	    #Net Stock greater than or eequal to 0 and no demand = No Action
		if debug: print "[Phase:5b]Runner item marked with No Action"
		sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
		sheet2.cell(row=rowNum, column=Comments).value = '>180 days of stock available' # JK Comment(col.59) 

#runners that need forecasted information for a decision
	if sheet.cell(row=rowNum, column=ABC).value == 'RUN' and sheet2.cell(row=rowNum, column=Action).value is None and sheet2.cell(row=rowNum, column=Action).value != 'No Action': 

		predictednet = returnvalue(rowNum)
		safety = sheet.cell(row=rowNum, column=SafetyStock).value

		if predictednet > safety:
			if debug: print "[Phase:5c]Runner item marked with No Action"
			sheet2.cell(row=rowNum, column=Action).value = 'No Action' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=Comments).value = 'Enought S/S to Cover demand' # JK Comment(col.59) 
		else:	
			if debug: print "[Phase:5d]Runner item marked with Order Action"
			sheet2.cell(row=rowNum, column=Action).value = 'Order' # Action(col.58) 'Note: We are updating the workbook with formulas from sheet2*'
			sheet2.cell(row=rowNum, column=OrderQty).value = getammount(safety,predictednet,rowNum) #Replen safety while ordering to meet predicted demand according to EBQ: bias = 0 default
			#1.25 bias applied to our runners; order that little bit more than usual.
			sheet2.cell(row=rowNum, column=Comments).value = 'S/S:'+str(safety)+' p.net:'+str(round(predictednet))+' (safety-predicted) = '+str(round(safety-predictednet)) # JK Comment(col.59) 


if debug: plt.show()

#remove graphs and dashboard
ws3 = wb2.get_sheet_by_name('Graphs')

if ws3 is not None:
	wb2.remove_sheet(ws3)



print('Saving purchase review...')
wb2.save('Purchase Review'+' '+time.strftime('%d-%m-%y')+'.xlsx') #save the modified workbook containing formulae
print("Done --- %sseconds ---" % (time.time() - start_time))
